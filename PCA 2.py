# Using image data as the labels

import pandas as pd
import openpyxl
from sklearn import preprocessing
from sklearn.decomposition import PCA
import numpy as np
import matplotlib.pyplot as plt



ua = []
ui = []
num_spots = []
total_coverage = []
avg_spot_size = []
spot_size_std = []
avg_circ = []
circ_std = []



wb = openpyxl.load_workbook(filename = "C:/Users/natha/OneDrive/Desktop/Summer 2023  Image analysis/Ua vs Ui "
                                       "Data/Final Data Raw.xlsx")
ws = wb.active # Get workbook active sheet object from active attribute

# extracts data from excel
for i in range(2, ws.max_row):
    ua.append(ws.cell(row=i, column=2).value)
    ui.append(ws.cell(row=i, column=3).value)
    num_spots.append(ws.cell(row=i, column=4).value)
    total_coverage.append(ws.cell(row=i, column=5).value)
    avg_spot_size.append(ws.cell(row=i, column=6).value)
    spot_size_std.append(ws.cell(row=i, column=7).value)
    avg_circ.append(ws.cell(row=i, column=8).value)
    circ_std.append(ws.cell(row=i, column=9).value)

Y = num_spots

ua_div_ui = []
for i in range(len(ui)):
    ua_div_ui.append(ua[i]/ui[i])

data = pd.DataFrame(index = Y)


names = ["ua","ui"]
categories = [ua,ui]

for j in range(len(categories)):
    data[names[j]] = categories[j]

print(data)
scaled_data = preprocessing.scale(data)

# print(scaled_data)

pca = PCA()

pca.fit(scaled_data)
pca_data = pca.transform(scaled_data)

#Scree Plot
per_var = np.round(pca.explained_variance_ratio_* 100, decimals=1)
labels = ['PC' + str(x) for x in range(1, len(per_var) + 1)]

plt.bar(x=range(1, len(per_var) + 1), height=per_var, tick_label=labels)
plt.ylabel('Percentage of Explained Variance')
plt.xlabel('Principal Component')
plt.title('Scree Plot')
plt.show()

# the following code makes a fancy looking plot using PC1 and PC2
pca_df = pd.DataFrame(pca_data, index= Y, columns = labels)

plt.scatter(pca_df.PC1, pca_df.PC2)
plt.title('My PCA Graph')
plt.xlabel('PC1 - {0}%'.format(per_var[0]))
plt.ylabel('PC2 - {0}%'.format(per_var[1]))

for sample in pca_df.index:
    plt.annotate(sample, (pca_df.PC1.loc[sample], pca_df.PC2.loc[sample]))

plt.show()

## get the name of the top 10 measurements (genes) that contribute
## most to pc1.
## first, get the loading scores
loading_scores = pd.Series(pca.components_[0], index=names)
## now sort the loading scores based on their magnitude
sorted_loading_scores = loading_scores.abs().sort_values(ascending=False)

# get the names of the top 10 genes
top_n = sorted_loading_scores[0:6].index.values

## print the gene names and their scores (and +/- sign)
print(loading_scores[top_n])