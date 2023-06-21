from mpl_toolkits import mplot3d
import matplotlib.pyplot as plt
import openpyxl
from matplotlib import cm
import numpy as np

ua = []
ui = []
num_spots = []
total_coverage = []
avg_spot_size = []
spot_size_std = []
avg_circ = []
circ_std = []



wb = openpyxl.load_workbook(filename = "C:/Users/natha/OneDrive/Desktop/Summer 2023  Image analysis/Ua vs Ui "
                                       "Data/Final Data Raw.xlsx")
ws = wb.active # Get workbook active sheet object from active attribute

# extracts data from excel
for i in range(2, ws.max_row):
    ua.append(ws.cell(row=i, column=2).value)
    ui.append(ws.cell(row=i, column=3).value)
    num_spots.append(ws.cell(row=i, column=4).value)
    total_coverage.append(ws.cell(row=i, column=5).value)
    avg_spot_size.append(ws.cell(row=i, column=6).value)
    spot_size_std.append(ws.cell(row=i, column=7).value)
    avg_circ.append(ws.cell(row=i, column=8).value)
    circ_std.append(ws.cell(row=i, column=9).value)


# fig1 = plt.figure()
# fig2 = plt.figure()
# # ax = plt.axes(projection='3d')
# ax1 = fig1.add_subplot(111, projection='3d')
# ax2 = fig2.add_subplot(111, projection='3d')

# z = [0]* len(num_spots)
# dx = [0.001] * len(ua)
# dy = [0.01] * len(ui)

# ax1.bar3d(ua, ui, z,dx,dy,num_spots)
# ax1.set_xlabel('Ua')
# ax1.set_ylabel('Ui')
# ax1.set_zlabel('Image Analysis Parameters')
#
# ax2.bar3d(ua, ui, z,dx,dy,total_coverage)
# ax2.set_xlabel('Ua')
# ax2.set_ylabel('Ui')
# ax2.set_zlabel('Image Analysis Parameters')



# ax.scatter3D(ua, ui, num_spots)
# ax.scatter3D(ua, ui, total_coverage)
# ax.scatter3D(ua,ui,avg_spot_size)
# ax.scatter3D(ua,ui,spot_size_std)
# ax.scatter3D(ua,ui,avg_circ)
# ax.scatter3D(ua,ui,circ_std)
#
# ax.set_xlabel('Ua')
# ax.set_ylabel('Ui')
# ax.set_zlabel('Image Analysis Parameters')


# Surface Plot
# Creating figure
# fig = plt.figure()
# ax = plt.axes(projection='3d')
#
# # Creating plot
# ax.plot_surface(np.array(ua), np.array(ui), np.array(num_spots),cmap=cm.coolwarm, linewidth=0, antialiased=False)
#
# # show plot
# plt.show()
#

