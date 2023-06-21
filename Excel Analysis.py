import openpyxl
import os
import numpy


# Note: The first row or column integer is 1, not 0.

directory = 'C:/Users/natha/OneDrive/Desktop/Summer 2023  Image analysis/Ua vs Ui Data/'

files = [] # list of the paths of all excel docs in the folder



# iterate over files in directory and add them to files
for filename in os.listdir(directory):
    f = directory + filename

    # checking if it is a file
    if os.path.isfile(f):
        files.append(f)


# iterate through excel files and do data analysis on them
for path in files:
    # open workbook and workbook object is created
    wb = openpyxl.load_workbook(filename = path)


    ws = wb.active # Get workbook active sheet object from active attribute

    num_spots = ws.max_row - 3 # get number of spots

    circ = []
    spot_size = []
    for i in range(4,num_spots+4): # adds spot sizes and circularity to arrays for all spots
        spot_size.append(ws.cell(row = i, column = 2).value)
        circ.append(ws.cell(row = i, column = 5).value)
    total_coverage = sum(spot_size)
    avg_spot_size = total_coverage/num_spots
    avg_circ = sum(circ)/num_spots
    spot_size_std = numpy.std(spot_size)
    circ_std = numpy.std(circ)

    print(path, ",",num_spots,",", total_coverage,",", avg_spot_size,",", spot_size_std,",", avg_circ,",", circ_std)








